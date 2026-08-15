#!/usr/bin/env python3
"""Fill my info into the internship xlsx."""
import openpyxl
import shutil
import os

src = "/opt/xiaonai/tmp_实习岗位.xlsx"
out = "/opt/xiaonai/2026年武汉市大学生实习实训岗位信息登记表_已填写.xlsx"

shutil.copy(src, out)

wb = openpyxl.load_workbook(out)
ws = wb.active

# Find row where 序号=1 and fill student info
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    seq_val = row[8].value  # 序号 column (index 8)
    if seq_val == 1:
        row[9].value = "小奈"       # 姓名
        row[10].value = "YOUR_COLLEGE"  # 学院
        row[11].value = "YOUR_MAJOR"  # 专业
        row[12].value = "大二"       # 年级班级
        row[14].value = "BOT_QQ_PLACEHOLDER" # 联系方式
        row[15].value = "是"         # 是否接收调剂
        break

wb.save(out)
print(f"OK: {out} ({os.path.getsize(out)} bytes)")
