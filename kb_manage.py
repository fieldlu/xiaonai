#!/usr/bin/env python3
"""Knowledge base management: CRUD + document import + search.
Usage:
  kb_manage.py list                          List all entries
  kb_manage.py view <topic>                  View entry content
  kb_manage.py add <topic> <content>         Add new entry
  kb_manage.py update <topic> <content>      Update (replace) existing entry
  kb_manage.py delete <topic>                Delete entry (fuzzy match)
  kb_manage.py search <keyword>              Search entries by keyword
  kb_manage.py import <filepath> [topic]     Import doc/txt/pdf/xlsx as entry
  kb_manage.py reindex                       Rebuild index.json from files
"""
import sys, os, json, glob, re, subprocess, tempfile, difflib
import jieba

KB = "/opt/xiaonai/data/knowledge"
INDEX = os.path.join(KB, "index.json")
BM25_FILE = os.path.join(KB, "bm25_index.pkl")
_BM25_IDX = None
_BM25_MTIME = 0
_BM25_STOP = {
    "一下","这个","那个","我们","你们","他们","咱们","现在","时候","什么","怎么",
    "如何","为什么","哪里","哪些","哪个","多少","几个","一个","一些","一点",
    "可以","不能","没有","还有","都是","不是","就是","因为","所以","如果","虽然",
    "但是","还是","已经","正在","即将","每个","这里","那里","请问","麻烦","帮忙",
    "谢谢","看看","了解","知道","问题","需要","应该","可能","比较","非常","特别",
    "真的","其实","然后","接着","最后","先","再","是不是","是不是","有没有",
}


# --- BM25 search (jieba) ---

def _bm25_build():
    """Build BM25 index from KB .md files using jieba tokenization."""
    import pickle, math
    from collections import Counter
    files = sorted(glob.glob(os.path.join(KB, "*.md")))
    doc_names, doc_tf, doc_len, doc_freq = [], [], [], Counter()
    for fp in files:
        if os.path.basename(fp) == "index.json":
            continue
        doc_names.append(fp)
        try:
            with open(fp, "r", encoding="utf-8", errors="ignore") as fh:
                text = fh.read()
        except Exception:
            text = ""
        toks = [t for t in jieba.cut(text) if len(t.strip()) >= 2 and t not in _BM25_STOP]
        tf = Counter(toks)
        doc_tf.append(tf)
        doc_len.append(len(toks))
        for t in set(toks):
            doc_freq[t] += 1
    n = len(doc_names)
    avgdl = (sum(doc_len) / n) if n else 0.0
    idf = {t: math.log((n - df + 0.5) / (df + 0.5) + 1.0) for t, df in doc_freq.items()}
    with open(BM25_FILE, "wb") as fh:
        pickle.dump({
            "doc_names": doc_names, "doc_tf": doc_tf, "doc_len": doc_len,
            "avgdl": avgdl, "idf": idf, "n": n, "k1": 1.5, "b": 0.75,
        }, fh)
    print(f"[bm25] Built index: {n} docs, {len(idf)} terms")


def _bm25_search(keyword, top_n=8):
    """BM25 keyword search. Returns [(topic, count, [snippet])], [] or None (no index)."""
    global _BM25_IDX, _BM25_MTIME
    import pickle
    if not os.path.exists(BM25_FILE):
        return None
    mt = os.path.getmtime(BM25_FILE)
    if _BM25_IDX is None or mt != _BM25_MTIME:
        with open(BM25_FILE, "rb") as fh:
            _BM25_IDX = pickle.load(fh)
        _BM25_MTIME = mt
    idx = _BM25_IDX
    q_toks = [t for t in jieba.cut(keyword) if len(t.strip()) >= 2 and t not in _BM25_STOP]
    if not q_toks:
        return None
    k1, b = idx["k1"], idx["b"]
    avgdl, n = idx["avgdl"], idx["n"]
    idf, doc_tf, doc_len, doc_names = idx["idf"], idx["doc_tf"], idx["doc_len"], idx["doc_names"]
    scores = []
    for i in range(n):
        tf, dl = doc_tf[i], doc_len[i]
        s = 0.0
        for t in q_toks:
            if t in idf:
                f = tf.get(t, 0)
                if f:
                    s += idf[t] * (f * (k1 + 1)) / (f + k1 * (1 - b + b * dl / avgdl))
        if s > 0:
            scores.append((doc_names[i], s))
    scores.sort(key=lambda x: -x[1])
    results = []
    kl = keyword.lower()
    for fp, sc in scores[:top_n]:
        try:
            with open(fp, "r", encoding="utf-8", errors="ignore") as fh:
                lines = fh.read().splitlines()
        except Exception:
            lines = []
        hit = next((l.strip()[:120] for l in lines if kl in l.lower() and l.strip()), "")
        if not hit:
            hit = next((l.strip()[:120] for l in lines if l.strip() and not l.startswith("#")), "")
        results.append((_topic_from_file(fp), max(1, int(sc)), [hit]))
    return results




# ─── Index management ───

def _load_index():
    if os.path.exists(INDEX):
        with open(INDEX, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                return set(data)
            return set(data.keys()) if isinstance(data, dict) else set()
    return set()

def _save_index(idx):
    os.makedirs(KB, exist_ok=True)
    with open(INDEX, "w", encoding="utf-8") as f:
        json.dump(sorted(idx), f, ensure_ascii=False, indent=2)

def _safe_name(topic):
    sanitized = re.sub(r'[\\/:*?"<>|]', "_", topic)
    return sanitized.strip() + ".md"

def _topic_from_file(fpath):
    """Extract topic title from file: use # Heading in file, fallback to filename."""
    try:
        with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.startswith("# "):
                    return line[2:].strip()
    except Exception:
        pass
    basename = os.path.basename(fpath)
    return basename.replace(".md", "")


# ─── Content helpers ───

def _extract_file_text(filepath):
    """Extract text from a file, supporting .docx, .txt, .pdf, .xlsx."""
    ext = os.path.splitext(filepath)[1].lower()
    if not os.path.exists(filepath):
        return None, f"File not found: {filepath}"

    if ext in (".txt", ".md"):
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            return f.read(), None

    elif ext == ".docx":
        try:
            from docx import Document
            doc = Document(filepath)
            text = "\n".join(p.text for p in doc.paragraphs)
            return text.strip(), None
        except Exception as e:
            # Fallback: pandoc
            try:
                result = subprocess.run(
                    ["pandoc", filepath, "-t", "plain"],
                    capture_output=True, text=True, timeout=15
                )
                if result.returncode == 0:
                    return result.stdout.strip(), None
            except:
                pass
            return None, f"Failed to read docx: {e}"

    elif ext == ".pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return text.strip(), None
        except Exception as e:
            return None, f"Failed to read PDF: {e}"

    elif ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    vals = [str(c) if c is not None else "" for c in row]
                    rows.append("\t".join(vals))
                parts.append(f"=== {sheet.title} ===\n" + "\n".join(rows))
            return "\n\n".join(parts).strip(), None
        except Exception as e:
            return None, f"Failed to read xlsx: {e}"

    else:
        return None, f"Unsupported format: {ext}"


# ─── Commands ───

def cmd_list():
    files = sorted(glob.glob(os.path.join(KB, "*.md")))
    if not files:
        print("[kb] Knowledge base is empty")
        return
    print(f"[kb] {len(files)} entries:")
    for f in files:
        size = os.path.getsize(f)
        topic = _topic_from_file(f)
        mtime = os.path.getmtime(f)
        from datetime import datetime
        dt = datetime.fromtimestamp(mtime).strftime("%m-%d %H:%M")
        label = "  "
        print(f"  {label} {topic}")
        print(f"      {size} bytes | updated {dt}")


def cmd_view(topic):
    idx = _load_index()
    fname = _safe_name(topic)
    fpath = os.path.join(KB, fname)
    if not os.path.exists(fpath):
        # Fuzzy match
        for f in glob.glob(os.path.join(KB, "*.md")):
            if topic.lower() in os.path.basename(f).lower():
                fpath = f
                break
    if not os.path.exists(fpath):
        print(f"[kb] Not found: {topic}")
        return
    with open(fpath, "r", encoding="utf-8") as f:
        content = f.read()
    # Print with header
    print(f"=== {os.path.basename(fpath).replace('.md','')} ===")
    print(content[:50000])
    if len(content) > 50000:
        print(f"\n... ({len(content) - 50000} more bytes)")


def cmd_add(topic, content):
    fname = _safe_name(topic)
    fpath = os.path.join(KB, fname)
    if os.path.exists(fpath):
        print(f"[kb] Entry already exists: {topic}")
        print(f"      Use 'update' to modify, or 'delete' first")
        return

    # Ensure content starts with a title
    if not content.startswith("# ") and not content.startswith("#"):
        content = f"# {topic}\n\n{content}"

    with open(fpath, "w", encoding="utf-8") as f:
        f.write(content)
    idx = _load_index()
    idx.add(topic)
    _save_index(idx)
    print(f"[kb] Added: {topic} ({len(content)} chars)")


def cmd_update(topic, content):
    fname = _safe_name(topic)
    fpath = os.path.join(KB, fname)
    if not os.path.exists(fpath):
        print(f"[kb] Not found: {topic}")
        print(f"      Use 'add' to create new entry")
        return

    if not content.startswith("# ") and not content.startswith("#"):
        content = f"# {topic}\n\n{content}"

    with open(fpath, "w", encoding="utf-8") as f:
        f.write(content)

    # Ensure index has this topic
    idx = _load_index()
    idx.add(topic)
    _save_index(idx)
    print(f"[kb] Updated: {topic} ({len(content)} chars)")


def cmd_delete(topic):
    idx = _load_index()
    fname = _safe_name(topic)
    fpath = os.path.join(KB, fname)
    if os.path.exists(fpath):
        os.remove(fpath)
        idx.discard(topic)
        _save_index(idx)
        print(f"[kb] Deleted: {topic}")
        return
    # Fuzzy match
    for f in glob.glob(os.path.join(KB, "*.md")):
        basename = os.path.basename(f)
        if topic.lower() in basename.lower():
            os.remove(f)
            old_topic = _topic_from_file(f)
            idx.discard(old_topic)
            _save_index(idx)
            print(f"[kb] Deleted: {basename}")
            return
    print(f"[kb] Not found: {topic}")


def _extract_shorter_keywords(text):
    """Extract shorter keyword candidates from text for search retry."""
    seen = set()
    candidates = []

    def add(word):
        w = word.strip()
        if w and len(w) >= 2 and w not in seen:
            seen.add(w)
            candidates.append(w)

    parts = re.split(r'[\s,.!?;:()\[\]{}<>/\\|@#$%^&*+=_\-"\u201c\u201d]+', text)
    for p in parts:
        if not p:
            continue
        if 2 <= len(p) <= 4:
            add(p)
        elif len(p) > 4:
            for length in range(4, 1, -1):
                for i in range(len(p) - length + 1):
                    seg = p[i:i+length]
                    if re.search(r'[\u4e00-\u9fff]', seg):
                        add(seg)

    if not candidates:
        for length in range(4, 1, -1):
            for i in range(len(text) - length + 1):
                seg = text[i:i+length]
                if re.search(r'[\u4e00-\u9fff]', seg):
                    add(seg)

    candidates.sort(key=lambda x: -len(x))
    final = []
    seen_content = set()
    for c in candidates:
        if c in seen_content:
            continue
        seen_content.add(c)
        final.append(c)
    return final


def _do_search(keyword):
    """Core KB search, returns list of (topic, count, snippets)."""
    bm = _bm25_search(keyword)
    if bm:
        return bm
    results = []
    for f in glob.glob(os.path.join(KB, "*.md")):
        if os.path.basename(f) == "index.json":
            continue
        with open(f, "r", encoding="utf-8", errors="ignore") as fh:
            file_content = fh.read()
        kw_lower = keyword.lower()
        if kw_lower in file_content.lower():
            lines = file_content.split(chr(10))
            topic = _topic_from_file(f)
            matches = [(i, l.strip()[:120]) for i, l in enumerate(lines) if kw_lower in l.lower()]
            results.append((topic, len(matches), matches[:3]))
    # Fuzzy complement: merge fuzzy results when exact match is sparse
    if len(results) < 5:
        fuzzy = _do_fuzzy_search(keyword)
        if fuzzy:
            seen_topics = {r[0] for r in results}
            max_exact = max((r[1] for r in results), default=0)
            fuzzy_base = max(1, max_exact) if results else 10
            for topic, score, snippets in fuzzy:
                if topic not in seen_topics:
                    seen_topics.add(topic)
                    fcount = max(1, min(int(score * 6), fuzzy_base))
                    if results:
                        fcount = min(fcount, max_exact)
                    results.append((topic, max(1, fcount), snippets))
    results.sort(key=lambda x: x[1], reverse=True)
    return results


def _char_ngrams(s, n=2):
    """Character n-grams for Chinese text."""
    s = s.lower()
    return {s[i:i+n] for i in range(len(s)-n+1)}


def _subseq_score(pattern, text):
    """Check if pattern chars appear as subsequence in text (ordered, not necessarily contiguous).
    Returns fraction of pattern matched (0-1).
    """
    if not pattern or not text:
        return 0.0
    pi = 0
    for ch in pattern:
        pi = text.find(ch, pi)
        if pi < 0:
            break
        pi += 1
    else:
        # All chars found in order!
        return 1.0
    return pi / max(len(pattern), 1)


def _do_fuzzy_search(keyword, threshold=0.33):
    """Fuzzy KB search: dual-track scoring.
    Track A: bigram presence + subsequence match (handles distant/shortened keywords).
    Track B: sliding window SequenceMatcher (handles typos/local).
    Track C: title boost.
    """
    kw = keyword.lower().strip()
    if len(kw) < 3:
        return []
    noise_chars = set("有哪哪些什么怎么如何为什么吗的呢了是不是还是或者因为所以")
    # Denoised main query
    kw_main = "".join(c for c in kw if c not in noise_chars)
    if len(kw_main) < 2:
        kw_main = kw
    main_chars = set(kw_main)
    main_bigrams = {kw_main[i:i+2] for i in range(len(kw_main)-1)} if len(kw_main) > 1 else set()
    kw_bigrams = {kw[i:i+2] for i in range(len(kw)-1)}
    win_size = max(len(kw) + 4, int(len(kw) * 2.5))

    results = []
    for f in glob.glob(os.path.join(KB, "*.md")):
        if os.path.basename(f) == "index.json":
            continue
        with open(f, "r", encoding="utf-8", errors="ignore") as fh:
            raw_content = fh.read()
        file_content = raw_content.lower()
        content_chars = set(file_content)

        # === Track A: multi-factor content presence ===
        # Factor 1: bigram overlap
        bg_score = 0.0
        if main_bigrams:
            bg_score = sum(1 for bg in main_bigrams if bg in file_content) / max(len(main_bigrams), 1)
        # Factor 2: character recall (fallback)
        char_hits = len(main_chars & content_chars)
        char_recall = char_hits / max(len(main_chars), 1) if main_chars else 0
        # Factor 3: subsequence match (handles abbreviations like 资环→资源与环境)
        subseq = _subseq_score(kw_main, file_content)
        # Combined presence: prefer subsequence (strong signal), then bigram, then char recall
        if subseq >= 0.8:
            presence_score = max(0.5, subseq)
        elif bg_score > 0.2:
            presence_score = max(bg_score, char_recall * 0.4)
        else:
            presence_score = char_recall * 0.4
        if presence_score < 0.1 or char_hits < 2:
            continue

        # === Track B: sliding window ===
        best_seq_raw = 0.0
        best_seq_clean = 0.0
        step = max(1, win_size // 2)
        limit = max(1, min(len(file_content) - win_size, 3000))
        for i in range(0, limit, step):
            window = file_content[i:i+win_size]
            r1 = difflib.SequenceMatcher(None, kw, window).ratio()
            if r1 > best_seq_raw:
                best_seq_raw = r1
            if len(kw_main) >= 3:
                r2 = difflib.SequenceMatcher(None, kw_main, window).ratio()
                if r2 > best_seq_clean:
                    best_seq_clean = r2
            if best_seq_raw >= 0.90:
                break

        # === Track C: title boost ===
        title_boost = 0.0
        first_lines = file_content.split(chr(10))[:10]
        for line in first_lines:
            if "# " in line:
                for bg in kw_bigrams:
                    if bg in line:
                        title_boost += 0.06
                if any(kw[:3] in line for _ in [1]):
                    title_boost += 0.04
                break
        title_boost = min(title_boost, 0.20)

        # === Combined score ===
        window_score = max(best_seq_raw, best_seq_clean)
        if window_score > 0:
            score = window_score * 0.45 + presence_score * 0.35 + title_boost
        else:
            score = presence_score * 0.5 + title_boost
        score = min(score + 0.10, 0.95)

        if score >= threshold:
            lines = file_content.split(chr(10))
            topic_line = ""
            for line in lines:
                if line.startswith("# "):
                    topic_line = line[2:].strip()
                    break
            if not topic_line:
                topic_line = os.path.splitext(os.path.basename(f))[0]
            snippet_lines = []
            seen_snippets = set()
            for i, line in enumerate(lines):
                if any(ch in line.lower() for ch in kw[:3]):
                    stripped = line.strip()[:120]
                    if stripped not in seen_snippets:
                        seen_snippets.add(stripped)
                        snippet_lines.append((i, stripped))
                        if len(snippet_lines) >= 2:
                            break
            results.append((topic_line, round(score, 3), snippet_lines))
    results.sort(key=lambda x: x[1], reverse=True)
    return results


def _print_search_results(keyword, results):
    """Print formatted search results."""
    print(f"[kb] Found '{keyword}' in {len(results)} files:")
    for topic, count, snippets in results:
        print(f"\n  {topic} ({count} matches)")
        for i, snippet in snippets:
            print(f"    L{i}: ...{snippet}...")


def cmd_search(keyword):
    """Search within KB file contents. Auto-retries with shorter keywords."""
    results = _do_search(keyword)
    if results:
        _print_search_results(keyword, results)
        return

    retries = _extract_shorter_keywords(keyword)
    tried = [keyword]
    for kw in retries:
        if kw in tried:
            continue
        tried.append(kw)
        results = _do_search(kw)
        if results:
            print(f"[kb] No results for: {keyword}")
            print(f"[kb] Auto-retrying with shorter keyword: {kw}")
            _print_search_results(kw, results)
            return

    print(f"[kb] No results for: {keyword}")

def cmd_import(filepath, topic=None):
    """Import a file as a KB entry."""
    if not os.path.exists(filepath):
        print(f"[kb] File not found: {filepath}")
        return

    text, error = _extract_file_text(filepath)
    if error:
        print(f"[kb] {error}")
        return
    if not text.strip():
        print(f"[kb] Empty content extracted")
        return

    # Auto-generate topic from filename if not provided
    if not topic:
        base = os.path.basename(filepath)
        topic = os.path.splitext(base)[0]
        # Clean up common suffixes
        topic = re.sub(r"[_\-]\d{8,}|\(?\d\)?$", "", topic).strip()

    # Check if entry already exists, offer update
    fname = _safe_name(topic)
    fpath = os.path.join(KB, fname)
    exists = os.path.exists(fpath)

    content = f"# {topic}\n\n{text.strip()}"

    with open(fpath, "w", encoding="utf-8") as f:
        f.write(content)

    idx = _load_index()
    idx.add(topic)
    _save_index(idx)

    action = "Updated" if exists else "Added"
    ext = os.path.splitext(filepath)[1].lower()
    print(f"[kb] {action}: {topic} (via {ext}, {len(content)} chars)")


def cmd_reindex():
    """Rebuild index.json from actual .md files on disk."""
    files = sorted(glob.glob(os.path.join(KB, "*.md")))
    topics = set()
    for f in files:
        topic = _topic_from_file(f)
        # Read first line for a better title
        with open(f, "r", encoding="utf-8", errors="ignore") as fh:
            first_line = fh.readline().strip()
        # Use the markdown heading if available
        if first_line.startswith("# "):
            title = first_line[2:].strip()
            topics.add(title)
        else:
            topics.add(topic)

    _save_index(topics)
    print(f"[kb] Reindexed: {len(topics)} entries from {len(files)} files")


# ─── CLI ───


def cmd_semantic(keyword):
    """Semantic search using TF-IDF char n-grams. Returns ranked results by relevance."""
    try:
        from kb_semantic import search
        search(keyword)
    except ImportError:
        print("[kb] Semantic search not available. Install numpy first")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    cmd = sys.argv[1]

    if cmd == "list":
        cmd_list()

    elif cmd == "view":
        if len(sys.argv) < 3:
            print("[kb] Usage: kb_manage.py view <topic>")
            sys.exit(1)
        cmd_view(sys.argv[2])

    elif cmd == "add":
        if len(sys.argv) < 4:
            print("[kb] Usage: kb_manage.py add <topic> <content>")
            print("      For longer content: kb_manage.py add <topic> @<filepath>")
            sys.exit(1)
        topic = sys.argv[2]
        content = sys.argv[3]
        if content.startswith("@"):
            fpath = content[1:]
            if os.path.exists(fpath):
                text, err = _extract_file_text(fpath)
                if err:
                    print(f"[kb] {err}")
                    sys.exit(1)
                content = text
            else:
                print(f"[kb] File not found: {fpath}")
                sys.exit(1)
        cmd_add(topic, content)

    elif cmd == "update":
        if len(sys.argv) < 4:
            print("[kb] Usage: kb_manage.py update <topic> <content>")
            print("      Or: kb_manage.py update <topic> @<filepath>")
            sys.exit(1)
        topic = sys.argv[2]
        content = sys.argv[3]
        if content.startswith("@"):
            fpath = content[1:]
            if os.path.exists(fpath):
                text, err = _extract_file_text(fpath)
                if err:
                    print(f"[kb] {err}")
                    sys.exit(1)
                content = text
            else:
                print(f"[kb] File not found: {fpath}")
                sys.exit(1)
        cmd_update(topic, content)

    elif cmd == "delete":
        if len(sys.argv) < 3:
            print("[kb] Usage: kb_manage.py delete <topic>")
            sys.exit(1)
        cmd_delete(sys.argv[2])

    elif cmd == "search":
        if len(sys.argv) < 3:
            print("[kb] Usage: kb_manage.py search <keyword>")
            sys.exit(1)
        cmd_search(sys.argv[2])
    elif cmd == "semantic":
        if len(sys.argv) < 3:
            print("[kb] Usage: kb_manage.py semantic <query>")
            sys.exit(1)
        cmd_semantic(sys.argv[2])

    elif cmd == "import":
        if len(sys.argv) < 3:
            print("[kb] Usage: kb_manage.py import <filepath> [topic]")
            sys.exit(1)
        filepath = sys.argv[2]
        topic = sys.argv[3] if len(sys.argv) > 3 else None
        cmd_import(filepath, topic)

    elif cmd == "reindex":
        cmd_reindex()

    else:
        print(f"[kb] Unknown command: {cmd}")
        print(__doc__)
